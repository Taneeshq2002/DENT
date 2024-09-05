import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
import os

# Initialize global variables
file_path = None
input_file_path = None
current_row = 2  # Start from the second row (assuming the first row has headers)

# Function to select an Excel file for reading
def select_input_file():
    global input_file_path, current_row
    input_file_path = filedialog.askopenfilename(
        title="Select Excel File to Read",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if input_file_path:
        current_row = 2  # Reset to the first data row
        populate_fields(input_file_path)
        input_file_label.config(text=f"Selected Input File: {os.path.basename(input_file_path)}")
    else:
        input_file_label.config(text="No input file selected")

# Function to populate fields from the selected input file
def populate_fields(file_path):
    global current_row
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Check if there's data to populate
        if ws[f'A{current_row}'].value is None:
            messagebox.showinfo("End of Records", "No more records to display.")
            return
        
        name_entry.delete(0, tk.END)
        name_entry.insert(0, ws[f'A{current_row}'].value)
        
        age_entry.delete(0, tk.END)
        age_entry.insert(0, ws[f'B{current_row}'].value)
        
        email_entry.delete(0, tk.END)
        email_entry.insert(0, ws[f'C{current_row}'].value)
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read the file: {e}")

# Function to select an Excel file for saving
def select_output_file():
    global file_path
    file_path = filedialog.askopenfilename(
        title="Select Excel File to Save",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if file_path:
        output_file_label.config(text=f"Selected Output File: {os.path.basename(file_path)}")
    else:
        output_file_label.config(text="No output file selected")

# Function to save data to the selected output Excel file
def save_to_excel():
    global file_path, current_row
    name = name_entry.get()
    age = age_entry.get()
    email = email_entry.get()

    if not name or not age or not email:
        messagebox.showwarning("Input Error", "All fields are required!")
        return

    if not file_path:
        messagebox.showwarning("File Error", "Please select an output Excel file first!")
        return

    if not os.path.exists(file_path):
        # If the file doesn't exist, create a new workbook and add the headers
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Email"])
    else:
        # Load the existing workbook
        wb = load_workbook(file_path)
        ws = wb.active

    # Append the data to the Excel sheet
    ws.append([name, age, email])

    # Save the workbook
    wb.save(file_path)

    # Clear the fields
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)

    messagebox.showinfo("Success", "Data saved to Excel successfully!")
    
    # Move to the next record
    current_row += 1
    populate_fields(input_file_path)

# Setting up the tkinter window
root = tk.Tk()
root.title("Data Entry Application")
root.geometry("400x400")
root.configure(bg="#f7f7f7")

# Styling options
label_font = ("Helvetica", 12, "bold")
entry_font = ("Helvetica", 12)
button_font = ("Helvetica", 12, "bold")
button_color = "#4CAF50"
button_text_color = "#ffffff"

# Title Label
title_label = tk.Label(root, text="Enter Your Details", font=("Helvetica", 16, "bold"), bg="#f7f7f7")
title_label.pack(pady=10)

# Frame for the form
form_frame = tk.Frame(root, bg="#f7f7f7")
form_frame.pack(padx=20, pady=10)

# Name Label and Entry
tk.Label(form_frame, text="Name:", font=label_font, bg="#f7f7f7").grid(row=0, column=0, sticky="e", padx=5, pady=5)
name_entry = tk.Entry(form_frame, font=entry_font, width=30, relief="solid", bd=1)
name_entry.grid(row=0, column=1, padx=5, pady=5)

# Age Label and Entry
tk.Label(form_frame, text="Age:", font=label_font, bg="#f7f7f7").grid(row=1, column=0, sticky="e", padx=5, pady=5)
age_entry = tk.Entry(form_frame, font=entry_font, width=30, relief="solid", bd=1)
age_entry.grid(row=1, column=1, padx=5, pady=5)

# Email Label and Entry
tk.Label(form_frame, text="Email:", font=label_font, bg="#f7f7f7").grid(row=2, column=0, sticky="e", padx=5, pady=5)
email_entry = tk.Entry(form_frame, font=entry_font, width=30, relief="solid", bd=1)
email_entry.grid(row=2, column=1, padx=5, pady=5)

# File Selection Button for Input File
input_file_button = tk.Button(root, text="Select Input File", font=button_font, bg=button_color, fg=button_text_color,
                              width=15, relief="flat", bd=0, command=select_input_file)
input_file_button.pack(pady=10)

# Label to show the selected input file
input_file_label = tk.Label(root, text="No input file selected", font=("Helvetica", 10), bg="#f7f7f7")
input_file_label.pack()

# File Selection Button for Output File
output_file_button = tk.Button(root, text="Select Output File", font=button_font, bg=button_color, fg=button_text_color,
                               width=15, relief="flat", bd=0, command=select_output_file)
output_file_button.pack(pady=10)

# Label to show the selected output file
output_file_label = tk.Label(root, text="No output file selected", font=("Helvetica", 10), bg="#f7f7f7")
output_file_label.pack()

# Save Button
save_button = tk.Button(root, text="Save to Excel", font=button_font, bg=button_color, fg=button_text_color,
                        width=15, relief="flat", bd=0, command=save_to_excel)
save_button.pack(pady=20)

# Start the GUI loop
root.mainloop()
